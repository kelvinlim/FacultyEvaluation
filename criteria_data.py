#! /usr/bin/env python3
"""
Shared loader/validator for faculty performance criteria.

Supports two input formats (auto-detected by default):
  1. YAML  — criteria_tables.yaml (original)
  2. Excel — criteria_input.xlsx  (user-editable workbook, one sheet per mission)

The public API is:
    load_tables(source=None)
        source: explicit Path to a .yaml or .xlsx file.
                If omitted, prefers Excel when it exists, else falls back to YAML.

    load_tables_from_yaml(path=None)
    load_tables_from_excel(path=None)
        Force a specific format.  path defaults to the standard filename.
"""
from __future__ import annotations

from functools import lru_cache
from pathlib import Path

import yaml

DEFAULT_YAML = Path(__file__).with_name("criteria_tables.yaml")
DEFAULT_EXCEL = Path(__file__).with_name("criteria_input.xlsx")


# ── helpers ──────────────────────────────────────────────────────────

def _error(source: str, message: str) -> ValueError:
    return ValueError(f"Invalid criteria data in {source}: {message}")


# ── YAML loader ──────────────────────────────────────────────────────

def load_tables_from_yaml(path: Path | None = None) -> dict:
    path = path or DEFAULT_YAML
    raw = yaml.safe_load(path.read_text(encoding="utf-8"))

    if not isinstance(raw, dict) or not raw:
        raise _error(path.name, "top-level structure must be a non-empty mapping")

    tables: dict = {}

    for mission_name, mission_data in raw.items():
        if not isinstance(mission_name, str) or not mission_name.strip():
            raise _error(path.name, "mission names must be non-empty strings")
        if not isinstance(mission_data, dict):
            raise _error(path.name, f"'{mission_name}' must map to an object")

        note = mission_data.get("note")
        groups = mission_data.get("groups")

        if not isinstance(note, str) or not note.strip():
            raise _error(path.name, f"'{mission_name}.note' must be a non-empty string")
        if not isinstance(groups, list) or not groups:
            raise _error(path.name, f"'{mission_name}.groups' must be a non-empty list")

        normalized_groups = []
        for group_idx, group in enumerate(groups):
            gpath = f"{mission_name}.groups[{group_idx}]"
            if not isinstance(group, dict):
                raise _error(path.name, f"'{gpath}' must be an object")

            label = group.get("label")
            rows = group.get("rows")
            if not isinstance(label, str) or not label.strip():
                raise _error(path.name, f"'{gpath}.label' must be a non-empty string")
            if not isinstance(rows, list) or not rows:
                raise _error(path.name, f"'{gpath}.rows' must be a non-empty list")

            normalized_rows = []
            for row_idx, row in enumerate(rows):
                row_path = f"{gpath}.rows[{row_idx}]"
                if not isinstance(row, dict):
                    raise _error(path.name, f"'{row_path}' must be an object")

                criterion = row.get("criterion")
                levels = row.get("levels")
                is_promotion = row.get("promotion_relevant")

                if not isinstance(criterion, str) or not criterion.strip():
                    raise _error(path.name, f"'{row_path}.criterion' must be a non-empty string")
                if not isinstance(levels, list) or len(levels) != 4 or not all(isinstance(x, str) for x in levels):
                    raise _error(path.name, f"'{row_path}.levels' must be a list of exactly 4 strings")
                if not isinstance(is_promotion, bool):
                    raise _error(path.name, f"'{row_path}.promotion_relevant' must be a boolean")

                normalized_rows.append((criterion, levels, is_promotion))

            normalized_groups.append((label, normalized_rows))

        tables[mission_name] = {
            "note": note,
            "groups": normalized_groups,
        }

    return tables


# ── Excel loader ─────────────────────────────────────────────────────

def load_tables_from_excel(path: Path | None = None) -> dict:
    """Read the editable input workbook and return normalized tables.

    Expected sheet layout (per mission area):
      Row 1: Note (merged across columns A–F)
      Row 2: Column headers (Criterion | Levels 1–4 | Promotion Relevant)
      Row 3+: Group label rows (merged, maroon fill) and data rows.

    A row is identified as a *group label* when columns B–E are empty
    (the label text lives in column A as a merged cell).
    """
    from openpyxl import load_workbook  # deferred import to keep YAML-only path lightweight

    path = path or DEFAULT_EXCEL
    wb = load_workbook(path, read_only=True, data_only=True)

    tables: dict = {}

    for ws in wb.worksheets:
        mission_name = ws.title
        source_tag = f"{path.name}!{mission_name}"

        # Row 1: note
        note = ws.cell(row=1, column=1).value
        if not isinstance(note, str) or not note.strip():
            raise _error(source_tag, "row 1 must contain the mission note text")

        # Row 2: headers (skip)
        # Row 3+: groups and data
        groups: list[tuple[str, list]] = []
        current_label: str | None = None
        current_rows: list = []

        for row in ws.iter_rows(min_row=3, max_col=6, values_only=True):
            col_a, col_b, col_c, col_d, col_e, col_f = (
                (row[i] if i < len(row) else None) for i in range(6)
            )

            # Detect group label row: col A has text, cols B–E are all empty
            if col_a and not any([col_b, col_c, col_d, col_e]):
                # Save previous group if any
                if current_label is not None:
                    if not current_rows:
                        raise _error(source_tag, f"group '{current_label}' has no data rows")
                    groups.append((current_label, current_rows))
                current_label = str(col_a).strip()
                current_rows = []
                continue

            # Skip fully empty rows
            if not col_a:
                continue

            # Data row
            criterion = str(col_a).strip()
            levels = []
            for val in (col_b, col_c, col_d, col_e):
                if val is None or str(val).strip() == "":
                    raise _error(source_tag, f"criterion '{criterion}' has an empty level cell")
                levels.append(str(val).strip())

            promo_raw = str(col_f).strip().lower() if col_f else ""
            if promo_raw in ("yes", "true", "1"):
                is_promotion = True
            elif promo_raw in ("no", "false", "0", ""):
                is_promotion = False
            else:
                raise _error(
                    source_tag,
                    f"criterion '{criterion}': Promotion Relevant must be Yes or No, got '{col_f}'",
                )

            current_rows.append((criterion, levels, is_promotion))

        # Save last group
        if current_label is not None:
            if not current_rows:
                raise _error(source_tag, f"group '{current_label}' has no data rows")
            groups.append((current_label, current_rows))

        if not groups:
            raise _error(source_tag, "sheet has no criterion groups")

        tables[mission_name] = {
            "note": note,
            "groups": groups,
        }

    wb.close()
    return tables


# ── Auto-detecting public API ────────────────────────────────────────

@lru_cache(maxsize=1)
def load_tables(source: Path | None = None) -> dict:
    """Load criteria tables, auto-detecting format from file extension.

    When *source* is None, prefers the Excel workbook if it exists,
    otherwise falls back to YAML.
    """
    if source is not None:
        if source.suffix in (".xlsx", ".xlsm"):
            return load_tables_from_excel(source)
        return load_tables_from_yaml(source)

    if DEFAULT_EXCEL.exists():
        return load_tables_from_excel(DEFAULT_EXCEL)
    return load_tables_from_yaml(DEFAULT_YAML)
