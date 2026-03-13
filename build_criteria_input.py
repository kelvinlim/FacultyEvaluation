#! /usr/bin/env python3
"""
Generate an editable Excel input workbook from criteria_tables.yaml.

Each mission area gets its own worksheet. Users can edit criteria text
and toggle promotion relevance via a Yes/No dropdown, then run the
build scripts which read this workbook as input.

Usage:
    python build_criteria_input.py                     # writes criteria_input.xlsx
    python build_criteria_input.py my_custom_name.xlsx # writes to custom filename
"""
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from criteria_data import load_tables_from_yaml

# ── Defaults ──
DEFAULT_OUTPUT = Path(__file__).with_name("criteria_input.xlsx")

# ── Styling (light, edit-friendly) ──
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

HEADER_FILL = PatternFill(start_color="7A0019", end_color="7A0019", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

GROUP_FILL = PatternFill(start_color="7A0019", end_color="7A0019", fill_type="solid")
GROUP_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=9)

NOTE_FONT = Font(name="Calibri", italic=True, size=9)

BODY_FONT = Font(name="Calibri", size=9)
PROMO_YES_FONT = Font(name="Calibri", size=9, bold=True, color="7A0019")

COL_HEADERS = [
    "Criterion",
    "1 – Unsatisfactory",
    "2 – Low Satisfactory",
    "3 – High Satisfactory",
    "4 – Outstanding",
    "Promotion\nRelevant",
]


def build_input_workbook(output_path: Path) -> None:
    tables = load_tables_from_yaml()
    wb = Workbook()

    for idx, (mission_name, data) in enumerate(tables.items()):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = mission_name

        # ── Note row (row 1, merged across all columns) ──
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        note_cell = ws.cell(row=1, column=1, value=data["note"])
        note_cell.font = NOTE_FONT
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[1].height = 36

        # ── Column headers (row 2) ──
        for col_idx, header in enumerate(COL_HEADERS, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = WRAP_CENTER
            cell.border = THIN_BORDER
        ws.row_dimensions[2].height = 30

        # ── Data validation for Promotion Relevant column ──
        promo_dv = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Invalid value",
            error="Please select Yes or No.",
        )
        promo_dv.prompt = "Select Yes or No"
        promo_dv.promptTitle = "Promotion Relevant?"
        ws.add_data_validation(promo_dv)

        # ── Data rows ──
        current_row = 3

        for group_label, rows in data["groups"]:
            # Group separator row
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=6,
            )
            sep_cell = ws.cell(row=current_row, column=1, value=f"  {group_label}")
            sep_cell.font = GROUP_FONT
            sep_cell.fill = GROUP_FILL
            sep_cell.alignment = Alignment(vertical="center")
            sep_cell.border = THIN_BORDER
            for c in range(2, 7):
                ws.cell(row=current_row, column=c).fill = GROUP_FILL
                ws.cell(row=current_row, column=c).border = THIN_BORDER
            ws.row_dimensions[current_row].height = 22
            current_row += 1

            for criterion, levels, is_promo in rows:
                ws.cell(row=current_row, column=1, value=criterion).font = Font(
                    name="Calibri", bold=True, size=9
                )
                ws.cell(row=current_row, column=1).alignment = WRAP_TOP
                ws.cell(row=current_row, column=1).border = THIN_BORDER

                for col_idx, val in enumerate(levels, start=2):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font = BODY_FONT
                    cell.alignment = WRAP_TOP
                    cell.border = THIN_BORDER

                promo_cell = ws.cell(
                    row=current_row, column=6, value="Yes" if is_promo else "No"
                )
                promo_cell.font = PROMO_YES_FONT if is_promo else BODY_FONT
                promo_cell.alignment = WRAP_CENTER
                promo_cell.border = THIN_BORDER
                promo_dv.add(promo_cell)

                current_row += 1

        # ── Column widths ──
        ws.column_dimensions["A"].width = 26
        for col_letter in ["B", "C", "D", "E"]:
            ws.column_dimensions[col_letter].width = 36
        ws.column_dimensions["F"].width = 12

        # Freeze below headers
        ws.freeze_panes = "A3"

    wb.save(output_path)
    print(f"Done! Saved editable input workbook to {output_path}")


if __name__ == "__main__":
    output = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_OUTPUT
    build_input_workbook(output)
