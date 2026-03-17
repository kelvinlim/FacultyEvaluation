#!/usr/bin/env python3
"""
Convert an Apple Numbers file to an Excel (.xlsx) file.

Usage:
    python convert_numbers_to_xlsx.py input.numbers
    python convert_numbers_to_xlsx.py input.numbers -o output.xlsx
"""
import argparse
import sys
from pathlib import Path

from numbers_parser import Document
from openpyxl import Workbook


def convert(input_path: Path, output_path: Path):
    doc = Document(str(input_path))
    wb = Workbook()
    wb.remove(wb.active)

    for sheet in doc.sheets:
        for table in sheet.tables:
            ws = wb.create_sheet(title=sheet.name[:31])
            for row_num, row in enumerate(
                table.iter_rows(min_row=0, max_row=table.num_rows - 1), 1
            ):
                for col_num, cell in enumerate(row, 1):
                    if cell.value is not None:
                        ws.cell(row=row_num, column=col_num, value=cell.value)

    if not wb.sheetnames:
        wb.create_sheet("Sheet1")

    wb.save(str(output_path))

    print(f"Saved {output_path}")
    for ws in wb.worksheets:
        print(f"  {ws.title}: {ws.max_row} rows x {ws.max_column} cols")


def main():
    parser = argparse.ArgumentParser(description="Convert Apple Numbers to Excel")
    parser.add_argument("input", type=Path, help="Input .numbers file")
    parser.add_argument("-o", "--output", type=Path, default=None, help="Output .xlsx file (default: same name with .xlsx extension)")
    args = parser.parse_args()

    if not args.input.exists():
        print(f"Error: file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    output = args.output or args.input.with_suffix(".xlsx")
    convert(args.input, output)


if __name__ == "__main__":
    main()
