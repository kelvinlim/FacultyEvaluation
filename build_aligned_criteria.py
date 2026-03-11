"""
Restructure Faculty Performance Criteria into horizontally-aligned tables.
Each row = one criterion topic, columns = rating levels 1-4.

Improvements:
- Color-coded rating columns (red → yellow → light green → green)
- Alternating row zebra striping
- Larger readable fonts, multi-page with repeating headers
- Separator rows between core duties and stretch/promotion criteria
- Landscape orientation
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.page import PageMargins
from criteria_data import load_tables

wb = Workbook()

# ── Styling ──
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
THICK_BOTTOM = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="medium"),
)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

# ── UMN Brand Colors ──
# Maroon: #7A0019, Gold: #FFCC33
# Extended palette: Light gold #FFDE7A, Dark maroon #5B0013, Light maroon tint #F0D5DC

# Column header fills & fonts — UMN Maroon
HEADER_FILL = PatternFill(start_color="7A0019", end_color="7A0019", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=9)

# Rating column background — neutral with warm UMN tint
ROW_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")         # white
ROW_FILL_ALT = PatternFill(start_color="FFF8EB", end_color="FFF8EB", fill_type="solid")    # very light gold

# Criterion column (col A) — light maroon tint
CRIT_FILL = PatternFill(start_color="F0D5DC", end_color="F0D5DC", fill_type="solid")        # light maroon
CRIT_FILL_ALT = PatternFill(start_color="E4BCC6", end_color="E4BCC6", fill_type="solid")    # slightly deeper maroon

# Promotion-highlight override — UMN Gold
PROMO_FILL = PatternFill(start_color="FFDE7A", end_color="FFDE7A", fill_type="solid")       # light gold
PROMO_FILL_ALT = PatternFill(start_color="FFCC33", end_color="FFCC33", fill_type="solid")   # UMN gold

# Separator row — UMN Maroon
SEP_FILL = PatternFill(start_color="7A0019", end_color="7A0019", fill_type="solid")
SEP_FONT = Font(name="Calibri", bold=True, size=8, color="FFFFFF")

# Body fonts
BODY_FONT = Font(name="Calibri", size=8.5)
BOLD_BODY_FONT = Font(name="Calibri", size=8.5, bold=True)
CRIT_FONT = Font(name="Calibri", bold=True, size=9)
CRIT_FONT_PROMO = Font(name="Calibri", bold=True, size=9, color="5B0013")  # dark maroon for promotion rows

# ── Data ──
tables = load_tables()

COL_HEADERS = ["Criterion", "1 – Unsatisfactory", "2 – Low Satisfactory", "3 – High Satisfactory", "4 – Outstanding"]

for idx, (sheet_name, data) in enumerate(tables.items()):
    ws = wb.active if idx == 0 else wb.create_sheet()
    ws.title = sheet_name

    # ── Page setup: landscape, 8.5x11, fit to 1 page WIDE but multi-page tall ──
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # allow multiple pages vertically
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.4, header=0.2, footer=0.2)
    ws.print_title_rows = "1:4"  # repeat title + headers on every page

    # ── Title row ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    title_cell = ws.cell(row=1, column=1, value=f"{sheet_name.upper()} PERFORMANCE CRITERIA")
    title_cell.font = Font(name="Calibri", bold=True, size=13, color="7A0019")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Note row ──
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    note_cell = ws.cell(row=2, column=1, value=data["note"])
    note_cell.font = Font(name="Calibri", italic=True, size=8.5)
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Row 3 = spacer
    ws.row_dimensions[3].height = 6

    # ── Column headers (row 4) ──
    for col_idx, header in enumerate(COL_HEADERS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[4].height = 28

    # ── Data rows ──
    current_row = 5
    data_row_count = 0  # for zebra striping (resets per group)

    for group_idx, (group_label, rows) in enumerate(data["groups"]):
        # Separator row
        if group_idx > 0:
            # Add a thin spacer row before separator
            ws.row_dimensions[current_row].height = 4
            current_row += 1

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        sep_cell = ws.cell(row=current_row, column=1, value=f"  {group_label}")
        sep_cell.font = SEP_FONT
        sep_cell.fill = SEP_FILL
        sep_cell.alignment = Alignment(vertical="center")
        sep_cell.border = THIN_BORDER
        # Fill all merged cells with border
        for c in range(2, 6):
            ws.cell(row=current_row, column=c).fill = SEP_FILL
            ws.cell(row=current_row, column=c).border = THIN_BORDER
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        data_row_count = 0

        for criterion, values, is_promo in rows:
            is_alt = (data_row_count % 2 == 1)

            # Column A: criterion name
            c = ws.cell(row=current_row, column=1, value=criterion)
            if is_promo:
                c.font = CRIT_FONT_PROMO
                c.fill = PROMO_FILL_ALT if is_alt else PROMO_FILL
            else:
                c.font = CRIT_FONT
                c.fill = CRIT_FILL_ALT if is_alt else CRIT_FILL
            c.alignment = WRAP_TOP
            c.border = THIN_BORDER

            # Columns B-E: rating values with color coding
            for col_idx, val in enumerate(values, start=2):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = BOLD_BODY_FONT if is_promo else BODY_FONT
                cell.alignment = WRAP_TOP
                cell.border = THIN_BORDER

                if is_promo:
                    cell.fill = PROMO_FILL_ALT if is_alt else PROMO_FILL
                else:
                    cell.fill = ROW_FILL_ALT if is_alt else ROW_FILL

            data_row_count += 1
            current_row += 1

    # ── Column widths for landscape (~10in printable) ──
    ws.column_dimensions["A"].width = 24
    for col_letter in ["B", "C", "D", "E"]:
        ws.column_dimensions[col_letter].width = 34

    # Freeze panes below header
    ws.freeze_panes = "A5"

output_path = Path(__file__).with_name("Faculty_Performance_Criteria_Aligned.xlsx")
wb.save(output_path)
print("Done! Saved to Faculty_Performance_Criteria_Aligned.xlsx")
